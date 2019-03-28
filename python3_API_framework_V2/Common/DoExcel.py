from openpyxl import load_workbook
import logging

class DoExcel:

    def __init__(self,excelpath):
        self.wb = load_workbook(excelpath)
        self.sh_case_data = self.wb["case_datas"]   #打开对应的标签页
        self.sh_prepare_data = self.wb["prepare_datas"]

    #从excel当中读取初始化数据，并对手机号码进行处理。
    def get_init_datas(self):
        init_datas = {}
        for index in range(2, self.sh_prepare_data.max_row + 1):
            key = self.sh_prepare_data.cell(row=index,column=1).value
            init_datas[key] = str(self.sh_prepare_data.cell(row=index,column=2).value)

        init_datas["${phone}"] = str(int(init_datas["${init_phone}"]) + 1)
        init_datas["${noReg_phone}"] = str(int(init_datas["${init_phone}"]) + 2)
        logging.info("初始化数据为：{0}".format(init_datas))
        #{"${init_phone}":"18612340000","${phone}":"18612340001","${phone}":"18612340002"}
        return init_datas

    def get_caseDatas_all(self):
        all_case_datas = []
        for index in range(2,self.sh_case_data.max_row+1):
            case_data = {}
            case_data["case_id"] = self.sh_case_data.cell(row=index,column=1).value
            case_data["api_name"] = self.sh_case_data.cell(row=index, column=4).value
            case_data["method"] = self.sh_case_data.cell(row=index, column=5).value
            case_data["url"] = self.sh_case_data.cell(row=index, column=6).value
            #case_data["request_data"] = self.sh_case_data.cell(row=index, column=7).value
            temp_case_data = self.sh_case_data.cell(row=index, column=7).value
            #获取初始值
            init_datas  = self.get_init_datas()
            # {"${init_phone}":"18612340000","${phone}":"18612340001","${phone}":"18612340002"}
            #temp_case_data = '{"mobilephone":"${init_phone}","pwd":"1234567890"}'
            if temp_case_data is not None and len(init_datas) > 0:
                for key,value in init_datas.items():
                    #如果找到了，则替换。
                    if temp_case_data.find(key) != -1:
                        temp_case_data = temp_case_data.replace(key,value)
            logging.info("初始化之后的请求数据为：\n{0}".format(temp_case_data))
            case_data["request_data"] = temp_case_data

            case_data["expected_data"] = self.sh_case_data.cell(row=index, column=8).value
            case_data["compare_type"] = self.sh_case_data.cell(row=index, column=9).value
            #判断本用例是否需要对响应结果进行解析，并获取其中的值。
            if self.sh_case_data.cell(row=index, column=10).value is not None:
                case_data["related_exp"] = self.sh_case_data.cell(row=index, column=10).value
            all_case_datas.append(case_data)
        return all_case_datas

    #更新初始化数据
    def update_init_data(self):
        init_data = self.sh_prepare_data.cell(row=2,column=2).value
        self.sh_prepare_data.cell(row=2, column=2).value = str(int(init_data) + 3)

    #保存数据
    def save_excelFile(self,excelpath):
        try:
            self.wb.save(excelpath)
        except Exception as e:
            logging.exception(e)
